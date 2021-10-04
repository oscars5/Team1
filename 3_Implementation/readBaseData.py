# __________________________IMPORTING PACKAGES_____________________________________________________
import openpyxl as xl
from openpyxl.chart import BarChart, PieChart, Reference
# import plotly
# import pandas as pd
# import numpy as np

# _________________________________READING INPUT SHEETS______________________________________________

# opening the source excel file
baseData = "Base_Data.xlsx"			# accessing base data sheet
wb1 = xl.load_workbook(baseData)
workSheet1 = wb1.worksheets[0]

keyData = "key.xlsx"				# accessing key sheet
wb2 = xl.load_workbook(keyData)
workSheet2 = wb2.worksheets[0]

cgpaData = "CGPA.xlsx"				# accessing cgpa calculator sheet
wb3 = xl.load_workbook(cgpaData)
workSheet3 = wb3.worksheets[0]


# opening the destination excel file
finalSheet = "final.xlsx"			# accessing final output sheet
wbFinal = xl.load_workbook(finalSheet)
workSheetFinal = wbFinal.active


# _______________________________NUMBER OF ROWS, COLUMNS_________________________________________________

# calculate total number of rows and columns in source excel file
mr = workSheet1.max_row
mc = workSheet1.max_column
print(mr, mc)

# ___________________________creating structure of base data and raw marks entry__________________________________

# copying the cell values from source excel file to destination excel file
x = 0
for l in range(6, mc+1):
	b = workSheet1.cell(row=1, column=l)
	workSheetFinal.cell(row=1, column=l + 2 * x + 2).value = b.value
	x = x+1

y = 1
for k in range(6, mc+1):
	workSheetFinal.cell(row=2, column=k + y, value='Score')
	workSheetFinal.cell(row=2, column=k + y + 1, value='Grades')
	workSheetFinal.cell(row=2, column=k + y + 2, value='Credits')
	y = y + 2
workSheetFinal.cell(row=2, column=k + y + 1, value='Average Score')
workSheetFinal.cell(row=2, column=6, value='CGPA')


# _______________________________RAW SCORE ENTRY,GRADES AND AVERAGE CALCULATION_______________________________________
for m in range(3, mr+2):
	z = 1
	eTotal = 0
	for n in range(6, mc+1):
		c = workSheet1.cell(row=m - 1, column=n)
		eTotal = eTotal + c.value
		workSheetFinal.cell(row=m, column=n + z).value = c.value
		if 90 <= c.value <= 100:
			workSheetFinal.cell(row=m, column=n + z+1).value = 'A+'
		elif 85 <= c.value < 90:
			workSheetFinal.cell(row=m, column=n + z+1).value = 'A'
		elif 80 <= c.value < 85:
			workSheetFinal.cell(row=m, column=n + z+1).value = 'A-'
		elif 75 <= c.value < 80:
			workSheetFinal.cell(row=m, column=n + z+1).value = 'B+'
		elif 70 <= c.value < 75:
			workSheetFinal.cell(row=m, column=n + z+1).value = 'B'
		elif 65 <= c.value < 70:
			workSheetFinal.cell(row=m, column=n + z+1).value = 'B-'
		elif 60 <= c.value < 65:
			workSheetFinal.cell(row=m, column=n + z+1).value = 'C+'
		elif 55 <= c.value < 60:
			workSheetFinal.cell(row=m, column=n + z+1).value = 'C'
		elif 50 <= c.value < 55:
			workSheetFinal.cell(row=m, column=n + z+1).value = 'C-'
		else:
			workSheetFinal.cell(row=m, column=n + z + 1).value = 'F'
		z = z+2
	workSheetFinal.cell(row=m, column=k + y + 1, value=eTotal / (mc - 5))

print(c.value)

for i in range(1, mr + 1):
	for j in range(1, 6):
		# reading cell value from source excel file
		c = workSheet1.cell(row=i, column=j)

		# writing the read value to destination excel file
		workSheetFinal.cell(row=i + 1, column=j).value = c.value

# ________________________________________________Credits Copying____________________________________________

for m1 in range(3, mr+2):
	z1 = 1
	x1 = 0

	for n1 in range(6, mc+1):
		c = workSheet2.cell(row=2, column=n1 + x1 - 3)
		x1 = x1 + 1

		workSheetFinal.cell(row=m1, column=n1 + z1 + 2).value = c.value
		z1 = z1+2

print(c.value)

# ____________________________________# TODO MIN, MAX, SD, TOP-10, BOTTOM-10___________________________________________


workSheetFinal.cell(row=mr + 3, column=5, value='AVERAGE')
workSheetFinal.cell(row=mr + 4, column=5, value='MAX')
workSheetFinal.cell(row=mr + 5, column=5, value='MIN')
workSheetFinal.cell(row=mr + 6, column=5, value='SD')
workSheetFinal.cell(row=mr + 7, column=5, value='KURT')

mrf = workSheetFinal.max_row
mcf = workSheetFinal.max_column
print(mrf, mcf)

for col in range(7, mcf, 3):
	max = 0
	min = 100
	mTotal = 0
	for row in range(3, mrf-5):
		c1 = workSheetFinal.cell(row=row, column=col)
		mTotal = mTotal + c1.value
		if max < c1.value:
			max = c1.value

		if min > c1.value:
			min = c1.value
		# print(c1.value)
	print(max)
	print(min)
	workSheetFinal.cell(row=row + 2, column=col, value=mTotal / (mr - 1))
	workSheetFinal.cell(row=row + 3, column=col, value=max)
	workSheetFinal.cell(row=row + 4, column=col, value=min)
print(mTotal)


# __________________________________# TODO HISTOGRAM AND BAR CHART____________________________________________________

# BAR CHART FOR AVERAGE SCORE

def bar_chart_avg():
	averageScore = Reference(workSheetFinal, min_col=7, min_row=22, max_col=mrf + 2, max_row=22)
	# Create object of BarChart class
	chart1 = BarChart()
	# adding data to the Bar chart object
	chart1.add_data(averageScore)
	# set the title of the chart
	chart1.title = " Bar Chart for Average "
	# set the title of the x-axis
	chart1.x_axis.title = " Modules "
	# set the title of the y-axis
	chart1.y_axis.title = " Average Scores "
	# add chart to the sheet.
	workSheetFinal.add_chart(chart1, "B28")


# BAR CHART FOR MAX SCORE

def bar_chart_max():
	maxScore = Reference(workSheetFinal, min_col=7, min_row=23, max_col=mrf + 2, max_row=23)
	chart2 = BarChart()
	chart2.add_data(maxScore)
	chart2.title = " Bar Chart for Maximum Score "
	chart2.x_axis.title = " Modules "
	chart2.y_axis.title = " Maximum Scores "
	workSheetFinal.add_chart(chart2, "K28")


# BAR CHART FOR MIN SCORE

def bar_chart_min():
	minScore = Reference(workSheetFinal, min_col=7, min_row=24, max_col=mrf + 2, max_row=24)
	chart3 = BarChart()
	chart3.add_data(minScore)
	chart3.title = " Bar Chart for Minimum Score "
	chart3.x_axis.title = " Modules "
	chart3.y_axis.title = " Minimum Scores "
	workSheetFinal.add_chart(chart3, "U28")


bar_chart_avg()
bar_chart_max()
bar_chart_min()

# _____________________________________Saving_________________________________________________________________________
# saving the destination excel file
wbFinal.save(str(finalSheet))


# ________________________________________# TODO TESTING PART_________________________________________________________
