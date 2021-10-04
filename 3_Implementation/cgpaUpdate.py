# ~~~~~~~~~~~~~~~~~~~~~~~PACKAGE IMPORTING~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
import openpyxl as xl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Color, PatternFill
from copy import copy

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~LOADING SHEETS~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

baseData = "Base_Data - Copy.xlsx"  # accessing base data sheet
wb1 = xl.load_workbook(baseData)
workSheet1 = wb1.active

keyData = "key.xlsx"                # accessing key sheet
wb2 = xl.load_workbook(keyData)
workSheet2 = wb2.worksheets[0]

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~NEW COLUMN STRUCTURE CREATION~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

add_border = Border(left=Side(style='thin'),    # setting border
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

workSheet1.insert_cols(6)        # insert a column

mr = workSheet1.max_row         # max row of base data
mc = workSheet1.max_column      # max column of base data

fill_color = PatternFill(patternType='solid', fgColor='FFFF00')     # choosing color
new_cell = workSheet1.cell(row=1, column=6, value="CGPA")
old_cell = workSheet1.cell(row=1, column=5)
new_cell.border = add_border
new_cell.alignment = Alignment(horizontal='center', vertical='center')
new_cell.font = copy(old_cell.font)
new_cell.fill = copy(old_cell.fill)
print(mr)
print(mc)

# ~~~~~~~~~~~~~~~~~~~~~~~CALCULATING CGPA UPDATING TO THE SAME SHEET~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

for i in range(2, mr + 1):
    workSheet1.cell(row=i, column=6).border = add_border
    workSheet1.cell(row=i, column=6).alignment = Alignment(horizontal='center')
    workSheet1.cell(row=i, column=6).fill = fill_color
    x = 0
    total_cg_score = 0
    total_credit = 0
    emp_cgpa = 0
    for data_col in range(7, mc + 1):
        module = workSheet2.cell(row=2, column=data_col - 4 + x-1)
        workSheet1.cell(row=1, column=data_col, value=module.value)
        credit = workSheet2.cell(row=2, column=data_col-4+x)
        total_credit = total_credit + credit.value
        x = x + 1
        score = workSheet1.cell(row=i, column=data_col)
        if 90 <= score.value <= 100:
            emp_cgpa = 4.5
        elif 85 <= score.value < 90:
            emp_cgpa = 4.25
        elif 80 <= score.value < 85:
            emp_cgpa = 4
        elif 75 <= score.value < 80:
            emp_cgpa = 3.75
        elif 70 <= score.value < 75:
            emp_cgpa = 3.5
        elif 65 <= score.value < 70:
            emp_cgpa = 3.25
        elif 60 <= score.value < 65:
            emp_cgpa = 3
        elif 55 <= score.value < 60:
            emp_cgpa = 2
        elif 50 <= score.value < 55:
            emp_cgpa = 1
        else:
            emp_cgpa = 0

        total_cg_score = total_cg_score + (emp_cgpa * credit.value)

        cgpa = total_cg_score/total_credit
        workSheet1.cell(row=i, column=6).value = round(cgpa, 2)
        # z1 = z1 + 2

print(credit.value)
print(score.value)
print(total_cg_score)

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~SAVING THE SHEET~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
wb1.save(str(baseData))
